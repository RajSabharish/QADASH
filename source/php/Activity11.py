__author__      = "varada vamsi"
__copyright__   = "Copyright 2017, NBN Buildco"

import openpyxl
from openpyxl import Workbook
import os.path
import re, sys
from collections import Counter


# ----------------------------------------------------------------------
class SiteDetails:
    try:
        def open_file(self, path):
            errorMemory = []
            errorMemory1 = []
            errorMemory2 = []
            """
            Open and read an Excel file
            """
            book = openpyxl.load_workbook(path)
            sheet = book.get_sheet_by_name('Site')
            sheet1 = book.get_sheet_by_name('Chassis')
            sheet2 = book.get_sheet_by_name('Patch')
            self.function1(sheet, errorMemory)
            self.function2(sheet1, errorMemory1)
            self.function3(sheet2, errorMemory2)
            textfile = open('activity11.txt', 'w')
            if len(errorMemory) > 0:
                templist = []
                print("Failed")
                for value, count in Counter(errorMemory).most_common():
                    templist.append(value +':'+ str(count))             
                status = "FAILED\n"
                desc = str(templist) + '\n'
                textfile.writelines(status)
                textfile.writelines(desc)
                templist[:] = []
            else:
                print("Pass")
                status = "PASS\n"
                desc = "NULL\n" + '\n'
                textfile.writelines(status)
                textfile.writelines(desc)



            if len(errorMemory1) > 0:
                print("Failed")
                for value, count in Counter(errorMemory1).most_common():
                    templist.append(value +':'+ str(count))
                status = "FAILED\n"
                desc = str(templist) + '\n'
                textfile.writelines(status)
                textfile.writelines(desc)
                templist[:] = []
            else:
                print("Pass")
                status = "PASS\n"
                desc = "NULL\n" + '\n'
                textfile.writelines(status)
                textfile.writelines(desc)

            if len(errorMemory2) > 0:
                print("Failed")
                for value, count in Counter(errorMemory2).most_common():
                    templist.append(value +':'+ str(count))
                status = "FAILED\n"
                desc = str(templist) + '\n'
                textfile.writelines(status)
                textfile.writelines(desc)
                templist[:] = []
            else:
                print("Pass")
                status = "PASS\n"
                desc = "NULL\n" + '\n'
                textfile.writelines(status)
                textfile.writelines(desc)

            if len(errorMemory) > 0 or len(errorMemory1) > 0 or len(errorMemory2) > 0:
                print("Failed")
                status = "FAILED\n"
                textfile.writelines(status)
            else:
                print("Success")
                status = "SUCCESS\n"
                textfile.writelines(status)
                
    except(Exception) as e:
        print ("File Improper")

    try:
        def function1(self, sheet, errorMemory):
            row_count = sheet.max_row
            for line1 in range(0, row_count - 1):
                if (sheet['B' + str(line1 + 2)].value) == None:
                    errorMemory.append('Site Code missing in Site:')
                if sheet['C' + str(line1 + 2)].value == None:
                    errorMemory.append('Designation missing in Site:')
                if sheet['J' + str(line1 + 2)].value == None:
                    errorMemory.append('Street Address missing in Site:')
                if sheet['K' + str(line1 + 2)].value == None:
                    errorMemory.append('Town missing in Site:')
                if sheet['L' + str(line1 + 2)].value == None:
                    errorMemory.append('State missing in Site:')
                if sheet['U' + str(line1 + 2)].value == None:
                    errorMemory.append('Object Status missing in Site:')
    except(Exception)as e:
        print ("File Improper")
    try:
        def function2(self, sheet, errorMemory1):
            row_count = sheet.max_row
            for line1 in range(0, row_count - 1):
                if (sheet['R' + str(line1 + 2)].value) != None:
                    val = str(sheet['M' + str(line1 + 2)].value.strip())
                    if not ((re.search('DSL', val)) or (re.search('CSD', val))):
                        errorMemory1.append('Equipment name missing in chasis:')
    except(Exception)as e:
        print ("File Improper")
    try:
        def function3(self, sheet, errorMemory2):
            row_count = sheet.max_row
            for line1 in range(0, row_count - 1):
                value = str(sheet['I' + str(line1 + 2)].value).strip()
                value1 = str(sheet['H' + str(line1 + 2)].value).strip()
                if re.search('SPLITTER',value):
                    if not (re.search('C', value1 )):
                        errorMemory2.append('C port missing in patch:')
    except(Exception)as e:
        print ("File Improper")

# ----------------------------------------------------------------------
if __name__ == "__main__":
    fibre = SiteDetails()
    if os.path.exists('activity11.txt'):
        os.remove('activity11.txt')

    #sam = '4WEP-20'
    sam = sys.argv[1]
    try:	
		destdir = 'C:\\QI_DATA\\' + sam
		files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
		for f in files:
			if re.search('Export_asset_capture_report', f):
				SiteFILE = 'C:\\QI_DATA\\' + sam + '\\' + f
				break
		path = SiteFILE + ""  # "Export_asset_capture_report_for_site_[FTTN_-_Compact_Sealed_DSLAM_(MicroNode)]_3ECH-_BACKUP.xlsx"
		if (path):
			fibre.open_file(path)
    except:
	    print ("File Improper")
