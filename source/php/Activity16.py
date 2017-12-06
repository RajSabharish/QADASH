from openpyxl import load_workbook
import os,re,sys, csv
class DSL:

    def function1(self,SAMNAME,dslname):
        wb = load_workbook(dslname)
        ws = wb['Site']
        No_Feild=0
        for designation in range(1,27):
            if ws.cell(row=1,column=designation).value=="Designation":
                No_Feild=1
                break
        for sitetype in range(1,27):
            if ws.cell(row=1,column=sitetype).value=="Site Type":
                No_Feild=1
                break

        ws = wb['Patch']
        for a_end in range(1,27):
            if ws.cell(row=1,column=a_end).value=="A End Chassis":
                No_Feild=1
                break
        for a_end_card in range(1,27):
            if ws.cell(row=1,column=a_end_card).value=="A End Card":
                No_Feild=1
                break
        for z_end in range(1,40):
            if ws.cell(row=1,column=z_end).value=="Z End Chassis":
                No_Feild=1
                break
        if No_Feild !=1 and designation < 27 and sitetype < 27 and a_end < 27 and z_end < 27 :
            print ("File Improper")
            
        ws = wb['Site']       
        fnocount=0
        fnoname=[]
        pdunottraced=[]
        num_rows=ws.max_row
        for i in range(2,num_rows+1):
                temp_fno=ws.cell(row=i,column=designation).value
                temp_csd=ws.cell(row=i,column=sitetype).value
                
                if re.search('FNO',temp_fno) and 'CSD' not in temp_csd:
                    fnocount=fnocount+1
                    fnoname.append(temp_fno)
        
        ws = wb['Patch']
        num_rows=ws.max_row
        
        for k in range(fnocount):
            flag=0
            for j in range(2,num_rows):
                a1=ws.cell(row=j,column=a_end).value
                b1=ws.cell(row=j,column=a_end_card).value
                c1=ws.cell(row=j,column=z_end).value
                if re.search("DSL",a1) and re.search((fnoname[k])[:10],a1) and re.search('NTC',b1) and re.search('PDU',c1):
                    flag=1
                    break
            if flag==0:
                pdunottraced.append(fnoname[k])

        textfile = open('activity16.txt', 'w')
        if len(pdunottraced) > 0:
            print("Failed")
            status = "Failed\n"
            desc=str(pdunottraced)+'\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        else:
            print ("Pass")
            status = "Pass\n"
            desc = 'NULL\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        if len(pdunottraced) > 0:
            print("Failed")
            status = "Failed\n"
            textfile.writelines(status)
        else:
            print ("Success")
            status = "Pass\n"
            textfile.writelines(status)
        
if __name__ == '__main__':
    os1 = DSL()
    if os.path.exists('activity16.txt'):
        os.remove('activity16.txt')
    sam = sys.argv[1]
    destdir = 'C:\\QI_DATA\\' + sam
    try:
        files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
        for f in files:
            if re.search('Export_asset_capture_report', f):
                dslfile = 'C:\\QI_DATA\\' + sam + '\\' + f
                break

        os1.function1(sam, dslfile)
    except:
        print("File Improper")
        
    
