__author__      = "varada vamsi"
__copyright__   = "Copyright 2017, NBN Buildco"

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os.path
import re,sys
from collections import Counter

# ----------------------------------------------------------------------
class ISP:
    def open_file(self,path):
        errorMemory = []
        var = ''
        """
        Open and read an Excel file
        """
        book = openpyxl.load_workbook(path)
        sheet = book.get_sheet_by_name('Site')
        sheet1 = book.get_sheet_by_name('Rack')
        sheet2= book.get_sheet_by_name('Chassis')
        sheet3 = book.get_sheet_by_name('Card')
        sheet4 = book.get_sheet_by_name('Cable')
        sheet5 = book.get_sheet_by_name('Patch')
        self.function1(sheet,errorMemory)
        self.function2(sheet1,errorMemory)
        self.function3(sheet2,errorMemory)
        self.function4(sheet3,errorMemory)
        self.function5(sheet4,errorMemory)
        self.function6(sheet5,errorMemory)
        textfile = open('activity17.txt', 'w')
        if len(errorMemory) > 1:
            templist = []
            print("Failed")
            for value, count in Counter(errorMemory).most_common():
                templist.append(value + '::' + str(count))
            print("Failed")
            status = "FAILED\n"
            desc = str(templist) + '\n'
            textfile.writelines(status)
            textfile.writelines(desc)
            textfile.writelines('FAILED')
        else:
            print("Pass")
            print("Success")
            status = "PASS\n"
            desc = "NULL\n" + '\n'
            textfile.writelines(status)
            textfile.writelines(desc)
            textfile.writelines('SUCCESS')



    try:
        def function1(self,sheet,errorMemory):
            row_count = sheet.max_row
            for line1 in range(0, row_count-1):
                if (sheet['C' + str(line1 + 2)].value) == None:
                    errorMemory.append('Designation missing in Site:')
                if sheet['U' + str(line1 + 2)].value == None:
                    errorMemory.append('Object Status missing in Site:')
                if sheet['V' + str(line1 + 2)].value == None:
                    errorMemory.append('Owner missing in Site:')
                #self.writedata(errorMemory,'A')
    except(Exception)as e:
        print ("File Improper")
    try:                                                                                               
        def function2(self,sheet,errorMemory):
            row_count = sheet.max_row
            for line1 in range(0, row_count-1):
                if (sheet['Q' + str(line1 + 2)].value) == None:
                    errorMemory.append('Project Name missing in Rack:')
                if sheet['R' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Number missing in Rack:')
                if sheet['T' + str(line1 + 2)].value == None:
                    errorMemory.append('Constructed Actual date missing in Rack:')
                if sheet['X' + str(line1 + 2)].value == None:
                    errorMemory.append('Object Status missing in Rack:')
                if sheet['Y' + str(line1 + 2)].value == None:
                    errorMemory.append('Owner missing in Rack:')
            #self.writedata(errorMemory,'B')
                                                                                                       
    except(Exception)as e:
        print ("File Improper")
    try:                                                                                                              
        def function3(self,sheet,errorMemory):
            row_count = sheet.max_row
            #print row_count                                                                                          
            for line1 in range(0, row_count-1):                                                                       
                if (sheet['M' + str(line1 + 2)].value) == None:
                    errorMemory.append('Name missing in Chasis:')
                if sheet['V' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Name missing in Chasis')
                if sheet['W' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Number missing in Chasis:')
                if sheet['Y' + str(line1 + 2)].value == None:
                    errorMemory.append('Construction actual date missing in Chasis')
                if sheet['AD' + str(line1 + 2)].value == None:
                    errorMemory.append('Owner missing in Chasis')
            #self.writedata(errorMemory,'C')
    except(Exception)as e:
        print ("File Improper")
    try:
        def function4(self,sheet,errorMemory):
            row_count = sheet.max_row
            #print row_count
            for line1 in range(0, row_count-1):
                if (sheet['C' + str(line1 + 2)].value) == None:
                    errorMemory.append('Designation Missing in Card')
                if sheet['W' + str(line1 + 2)].value == None:
                    errorMemory.append('Rack Missing in Card')
                if sheet['T' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Name Missing in Card')
                if sheet['U' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Number Missing in Card')
                if sheet['AA' + str(line1 + 2)].value == None:
                    errorMemory.append('Object Status Missing in Card')
                if sheet['AB' + str(line1 + 2)].value == None:
                    errorMemory.append('Owner Missing in Card')
                #self.writedata(errorMemory,'D')

    except(Exception)as e:
        print ("File Improper")
    try:                                                                                                          
        def function5(self,sheet,errorMemory):
            row_count = sheet.max_row
            #print row_count                                                                                      
            for line1 in range(0, row_count-1):                                                                   
                if (sheet['I' + str(line1 + 2)].value) == None:
                    errorMemory.append('Cable Name Missing in Cable')
                if sheet['J' + str(line1 + 2)].value == None:
                    errorMemory.append('Construction Actual Date Missing in Cable')
                if sheet['N' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Name Missing in Cable')
                if sheet['O' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Number Missing in Cable')
                if sheet['S' + str(line1 + 2)].value == None:
                    errorMemory.append('Object Status Missing in Cable')
                if sheet['T' + str(line1 + 2)].value == None:
                    errorMemory.append('Owner Missing in Cable')
                #self.writedata(errorMemory,'E')
                                                                                                                      
    except(Exception)as e:
        print ("File Improper")

    try:                                                                                                          
        def function6(self,sheet,errorMemory):
            row_count = sheet.max_row
            #print row_count                                                                                      
            for line1 in range(0, row_count-1):                                                                   
                if (sheet['J' + str(line1 + 2)].value) == None:
                    errorMemory.append('Name Missing in Patch')
                if sheet['K' + str(line1 + 2)].value == None:
                    errorMemory.append('Construction actual date Missing in Patch')
                if sheet['M' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Name Missing in Patch')
                if sheet['N' + str(line1 + 2)].value == None:
                    errorMemory.append('Project Number Missing in Patch')
                if sheet['Q' + str(line1 + 2)].value == None:
                    errorMemory.append('Object Status Missing in Patch')
                #self.writedata(errorMemory,'F')

    except(Exception)as e:
        print ("File Improper")

    try:
        def writedata(self,list,column):
            dest_filename = 'error.xlsx'
            if os.path.isfile(dest_filename):
                wb = load_workbook(filename=dest_filename)
                ws1 = wb.active
                ws1.title = "errorbook"
                for i in range(0, len(list)):
                    ws1[column + str(i + 2)] = list[i]
                wb.save(dest_filename)
    except(Exception)as e:
        print ("File Improper")

# ----------------------------------------------------------------------
if __name__ == "__main__":
    fibre=ISP()
    if os.path.exists('activity17.txt'):
        os.remove('activity17.txt')
    sam = sys.argv[1]
    #sam = '4WEP-20'
    destdir = 'C:\\QI_DATA\\' + sam
    try:

        files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
        for f in files:
            if re.search('Export_asset_capture_report', f):
                ISPFILE = 'C:\\QI_DATA\\' + sam + '\\' + f
                break
        
        path = ISPFILE+"" #"Export_asset_capture_report_for_site_[FTTN_-_Compact_Sealed_DSLAM_(MicroNode)]_3ECH-_BACKUP.xlsx"
        if (path):
            fibre.open_file(path)
    except:
            print("File Improper")
