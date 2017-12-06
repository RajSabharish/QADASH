from openpyxl import *
import re,sys,os

class Activity15():
    def activity_15(self,source,files,sam):
        fnolist=[]
        Rackfnolist=[]
        catfnolist=[]
        wb=load_workbook(source)
        ws=wb['Rack']
        rowcount=ws.max_row

        for row in range(2,rowcount+1):
            fno=ws["C" + str(row)].value
            if fno is not None:
                if re.search("FNO",fno):
                    fnolist.append(fno)

        fnolist=list(set(fnolist))
        comfno=[]
        for row in range(2,rowcount+1):
            fno=ws["C" + str(row)].value

            if fno is not None and re.search("FNO",fno):
                if ws["F"+str(row)].value != "GD.1.1":
                    Rackfnolist.append(fno+'Racklocation')

            if fno is not None and re.search("FNO",fno) and ws["D"+str(row)].value == "NBN_FTTN_CSD":
                if ws["E"+str(row)].value != "1209280":
                    catfnolist.append(fno+'CSDCatNum')

            if fno is not None and re.search("FNO",fno) and ws["D"+str(row)].value == "NBN_FTTN" and (ws["G"+str(row)].value) != "COMMSCOPE":
                if ws["E"+str(row)].value != "1008461":
                    catfnolist.append(fno+'NokiaCatNum')

            if fno is not None and re.search("FNO",fno) and ws["D"+str(row)].value == "NBN_FTTN" and (ws["G"+str(row)].value) == "COMMSCOPE":
                comfno.append(fno)
                if ws["E"+str(row)].value != "1610240":
                    catfnolist.append(fno+'COMSCOPECatNum')



        #print(Rackfnolist,catfnolist)


        logicalname=[]
        logicalname1=[]
        ws = wb['Chassis']
        rowcount = ws.max_row

        for fno in fnolist:
            sam1 = sam + fno[7:10]
            filefound=0
            for f in files:
                if re.search('iPACT', f) and re.search(sam1, f):
                    IPACT = 'C:\\QI_DATA\\' + sam + '\\' + f
                    filefound=1
                    break

            if filefound == 1:
                wb1=load_workbook(IPACT)
                ws1=wb1['NT']
                rowcount1=ws.max_row
                for i in range(2,rowcount1+1):
                    logicalname.append(ws1["A"+str(i)].value)
            flag=0
            for i in range(2,rowcount + 1):
                if ws["M" + str(i)].value is not None and re.search("CSD-0001", ws["M" + str(i)].value) and fno == ws["C" + str(i)].value:
                    flag=1
                    break
            lname = ws["R" + str(i)].value
            if ((lname is None) or (lname not in logicalname)) and flag == 1 and filefound == 1:
                logicalname1.append(fno+'CSDFNOLogicalNameIncorrect')
            if ((lname is None) or (lname not in logicalname)) and flag == 1 and filefound == 0:
                logicalname1.append(fno+'CSDFNOLogicalNameIpactMissing')

            flag=0
            for i in range(2, rowcount + 1):
                if ws["M" + str(i)].value is not None and re.search("DSL-0001", ws["M" + str(i)].value) and fno == ws["C" + str(i)].value:
                    flag = 1
                    break
            lname = ws["R" + str(i)].value
            if ((lname is None) or (lname not in logicalname)) and flag == 1:
                logicalname1.append(fno + 'COM/Nokia_FNOLogicalName')

            flag=0
            Pysub=[]
            for i in range(2, rowcount + 1):
                if ws["M" + str(i)].value is not None and ( re.search("PDU-0001", ws["M" + str(i)].value) or re.search("PSY-0001", ws["M" + str(i)].value) or re.search("DSL-0001", ws["M" + str(i)].value)) and fno == ws["C" + str(i)].value:
                    flag = 1
                    break
            pname = ws["F" + str(i)].value
            #print(pname,fno)
            if (pname is None or pname != "GD.1.1.2" or pname != "GD.1.1.3" or pname != "GD.1.1.4") and flag == 1:
                #print(pname,fno)
                Pysub.append(fno + 'pysub/Nokia-COM_FNO1')

            #flag=0
            #for i in range(2, rowcount + 1):
            #    if ws["M" + str(i)].value is not None and re.search("PSY-0001", ws["M" + str(i)].value) and fno == ws["C" + str(i)].value:
            #        flag = 1
            #        break
            #pname = ws["F" + str(i)].value
            #if (pname is None or pname != "GD.1.1.3") and flag == 1:
            #    Pysub.append(fno + 'pysub/Nokia-COM_FNO2')
            #
            #flag=0
            #for i in range(2, rowcount + 1):
            #    if ws["M" + str(i)].value is not None and re.search("DSL-0001", ws["M" + str(i)].value) and fno == ws["C" + str(i)].value:
            #        flag = 1
            #        break
            #pname = ws["F" + str(i)].value
            #if (pname is None or pname != "GD.1.1.4") and flag == 1:
            #    Pysub.append(fno + 'pysub/Nokia-COM_FNO3')


            for i in range(2, rowcount + 1):
                if ws["M" + str(i)].value is not None and re.search("BAT-0001", ws["M" + str(i)].value) and fno == ws["C" + str(i)].value:
                    Pysub.append(fno + 'BAT/Nokia-COM_FNO')
                    break

        #print(logicalname1,Pysub)
        comscopebattery=[]
        for fno in comfno:
            ws2 = wb['Card']
            rowcount2 = ws.max_row
            vlccount = 0
            for i in range(2, rowcount2 + 1):
                tempvlc = ws2["N" + str(i)].value
                if tempvlc is not None and re.search("VLC", tempvlc) and re.search(fno,ws2["C" + str(i)].value) :
                        vlccount = vlccount + 1

            batcount = 0
            ws = wb['Chassis']
            for i in range(2, rowcount + 1):
                tempbat = ws["M" + str(i)].value
                if tempbat is not None and re.search("BAT", tempbat) and re.search(fno,ws["C" + str(i)].value):
                        batcount = batcount + 1
            #print(vlccount,batcount)
            if vlccount == 4  and batcount != 4:
                    #print(fno,"case 1 Failed")
                    comscopebattery.append(fno+"comscopebattery")
            if vlccount > 4 and batcount != 8:
                    #print(fno,"Case 2 Failed")
                    comscopebattery.append(fno+"comscopebattery")
        textfile = open('activity15.txt', 'w')
        if len(Rackfnolist) > 0 or len(logicalname1) > 0 or len(Pysub) > 0 :
            print("Failed")
            textfile.writelines("Failed\n")
            textfile.writelines(str(Rackfnolist+logicalname1+Pysub)+"\n")
        else:
            print("Pass")
            textfile.writelines("Pass\n")
            textfile.writelines("NULL\n")
        if len(catfnolist) > 0:
            print("Failed")
            textfile.writelines("Failed\n")
            textfile.writelines(str(catfnolist)+"\n")
        else:
            print("Pass")
            textfile.writelines("Pass\n")
            textfile.writelines("NULL\n")
        if len(comscopebattery) > 0:
            print("Failed")
            textfile.writelines("Failed\n")
            textfile.writelines(str(comscopebattery)+"\n")
        else:
            print("Pass")
            textfile.writelines("Pass\n")
            textfile.writelines("NULL\n")
        if len(Rackfnolist) > 0 or len(logicalname1) > 0 or len(Pysub) > 0 or comscopebattery == "Failed" or len(catfnolist) > 0 :
            print("Failed")
            textfile.writelines("Failed\n")
        else:
            print("Success")
            textfile.writelines("Success\n")


if __name__=="__main__":
    ob=Activity15()
    if os.path.exists('activity15.txt'):
        os.remove('activity15.txt')
    sam=sys.argv[1]
    destdir = 'C:\\QI_DATA\\' + sam
    try:
        files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
        for f in files:
            if re.search('Export_asset_capture_report_for_site', f):
                sourcefile = 'C:\\QI_DATA\\' + sam + '\\' + f
                break
        ob.activity_15(sourcefile,files,sam)
    except:
        print("File Improper")
        sys.exit()
    
        

            
