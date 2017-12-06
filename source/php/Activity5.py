import re, os , sys , csv
from openpyxl import load_workbook
class CSSTieUp:
    def CSSValidation(self,sam,ospextract,cssosp,csscopper):
        
        
        ospcss={}
        adacss1=[]
        v1_failed=[]
        v2_failed=[]
        v4_failed=[]
        v1,v2,v3,v4=('','','','')
        mantype=[]
        
        
        wb=load_workbook(cssosp)
        ws=wb['Plant - Cables']
        rowcount=ws.max_row
        
        for cable in range(1,27):
            a1=ws.cell(row=1,column=cable).value            
            if "Cable Name" == a1:
                nofield1=1
                break
        for mtype in range(1,27):
            b1=ws.cell(row=1,column=mtype).value            
            if "Manufacturer Type Description" == b1:
                nofield2=1
                break
            
        if (nofield1 != 1 or nofield2 != 1) and cable < 26 and mtype < 26:
            print('File Improper')
            sys.exit()
            
        for j in range(2,rowcount+1):
            c1=ws.cell(row=j,column=cable).value
            d1=ws.cell(row=j,column=mtype).value
            if c1 is not None:
                if re.search("CSS",c1):
                    ospcss.update({c1:d1})
                      
                 
        
        wb1=load_workbook(ospextract)
        ws1=wb1['Plant - Cables']
        rowcount=ws1.max_row
        for i in range(2,rowcount +  1):
            if re.search('CSS',ws1["B"+str(i)].value):
                        adacss1.append(ws1["B"+str(i)].value)
                        mantype.append(ws1["D"+str(i)].value)

        adacss2=list(set(adacss1))
        keys=list(ospcss.keys())
        values=list(ospcss.values())
        for i in range(len(adacss2)):
            for j in range(len(keys)):
                if adacss2[i]==keys[j]:
                    pairtype1=re.search('\d{2,3}',mantype[i])
                    pairtype2=re.search('\d{2,3}',values[j])
                    break
            if (adacss1.count(adacss2[i]) > 1 or pairtype1.group(0)!= pairtype2.group(0)) and (j < len(keys)):
                v1="Failed"
                v1_failed.append(adacss2[i])
            
                    
        for k in range(len(adacss2)):
            if re.search('^\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}$',adacss2[k]) is None:
              v2="Failed"
              v2_failed.append(adacss2[k])
              
        csscable=[]
        with open(csscopper,'r') as csvfile:
            reader=csv.reader(csvfile)
            for row in reader:
                if re.search('CSS',row[1]):
                    csscable.append(row)
        for css in range(len(adacss2)):
            flag=0
            for row1 in csscable:
                if adacss2[css]==row1[1]:
                    
                    paircount=re.search('\d{2,3}',mantype[css])
                    paircount=paircount.group(0)
                    
                    coretype=re.search('[A-Z]+\/[A-Z]+',mantype[css])
                    coretype=coretype.group(0)
                    
                    diameter=re.search('\d\.\d+',mantype[css])
                    diameter=diameter.group(0)
                    
                    temp=row1[40]+'/'+row1[39]
                    
                    if row1[33] != paircount or float(row1[34]) != float(diameter) or temp != coretype:
                        v4="Failed"
                        v4_failed.append(adacss2[css])                                                   

        textfile = open('activity5.txt', 'w')
        if v1=='Failed':
            print('Failed')
            status = "Failed\n"
            desc=str(v1_failed)+'\n'
            textfile.writelines(status)
            textfile.writelines(desc)


        else:
            print('Pass')
            status = "Pass\n"
            desc = 'NULL\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        if v2=='Failed':
            print('Failed')
            status = "Failed\n"
            desc = str(v2_failed)+'\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        else:
            print('Pass')
            status = "Pass\n"
            desc = 'NULL\n'
            textfile.writelines(status)
            textfile.writelines(desc)

        if v4=='Failed':
            print('Failed')
            #print(v4_failed)
            status = "Failed\n"
            desc = str(v4_failed)+'\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        else:
            print('Pass')
            status = "Pass\n"
            desc = 'NULL\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        if v1=='Failed' or v2=='Failed' or v4=='Failed' :
            print('Failed')
            status = "Failed\n"
            textfile.writelines(status)
        else:
            print('Success')
            status = "Success\n"
            textfile.writelines(status)


if __name__=='__main__':
    cssob = CSSTieUp()
    if os.path.exists('activity5.txt'):
        os.remove('activity5.txt')
    sam=sys.argv[1]
    destdir = 'C:\\QI_DATA\\' + sam
    try:
        files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
        for f in files:
            if re.search(sam+'-OSP', f):
                cssosp  = 'C:\\QI_DATA\\' + sam + '\\' + f
                break
        for f in files:
            if re.search('OSP_Editor_Export_NBN', f):
                ospextract = 'C:\\QI_DATA\\' + sam + '\\' + f
                break

        files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
        for f in files:
            if re.search(sam+'_list_copper_cable', f):
                csscopper  = 'C:\\QI_DATA\\' + sam + '\\' + f
                break

        cssob.CSSValidation(sam, ospextract, cssosp,csscopper)
    except:
        print("File Improper")
        sys.exit()

            



