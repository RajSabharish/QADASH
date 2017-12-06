from openpyxl import load_workbook
import os,re,sys, csv

class OSP:
     
    def OSPValidation(self,osp):
       
        wb = load_workbook(osp)
        listoffields1=[]
        listoffields2=[]
        listoffields3=[]
        listoffields4=[]
        listoffields5=[]
        listoffields6=[]
        listoffields7=[]
        listoffields8=[]
        d1={}
        d2={}
        ## Find Blanks
        v1=[]
        v2=[]
        for sheet in wb.worksheets:
            flag=0
            ws = wb[sheet.title]
            if ws.title=="Civil - NBN Pits":
                
                num_rows=ws.max_row
                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['D'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields1.append("Project Name:"+str(pcount))              

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['E'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields1.append("Project Number:"+str(pcount)) 

                pcount=0         
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['F'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields1.append("Owner Name:"+str(pcount))

                pcount=0    
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['H'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields1.append("Life Cycle Status:" + str(pcount))
                                
                pcount=0           
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['i'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields1.append("Constructed Actual Date:"+str(pcount))

                      
                
                                

            elif ws.title=="Civil - NBN Trenches":
                num_rows=ws.max_row
                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['H'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields2.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['I'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields2.append("Life Cycle Status:"+str(pcount))
                    
                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['J'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields2.append("Constructed Actual Date:"+str(pcount))

                
                


            elif ws.title=="Civil - NBN Ducts":
                num_rows=ws.max_row

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['F'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields3.append("Project Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['G'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields3.append("Project Number:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['H'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields3.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['I'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields3.append("Life Cycle Status:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['J'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields3.append("Constructed Actual Date:"+str(pcount))
                
                          



                                
            elif ws.title=="Plant - Equipment": 
                num_rows=ws.max_row
                
                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['E'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields4.append("Project Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['F'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields4.append("Project Number:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['G'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields4.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['H'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields4.append("Life Cycle Status:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['I'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields4.append("Constructed Actual Date:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a) :
                            if (ws['L'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields4.append("Batch Number:"+str(pcount))

               
            
                            
            elif ws.title=="Plant - Cables": 
                num_rows=ws.max_row

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['H'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields5.append("Project Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['I'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields5.append("Project Number:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['J'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields5.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['L'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields5.append("Life Cycle Status:"+str(pcount))

                
                pcount=0
                for i in range(2,num_rows+1):        
                    a=ws['B'+str(i)].value
                    if a is not None:
                        if len(a)==18 and re.search('\d{1}\w{3}-\d{2}-\d{2}-\w{3}-\d{3}',a):
                            if (ws['N'+str(i)].value) is None:
                                pcount = pcount + 1
                if pcount > 0 :
                    listoffields5.append("Constructed Actual Date:"+str(pcount))

             
            elif ws.title=="Civil - Telstra Trenches": 
                num_rows=ws.max_row
                pcount=0
                for i in range(2,num_rows+1):
                    a=ws['B'+str(i)].value
                    if a is not None and (a.isnumeric()):
                        if (ws['F'+str(i)].value) != "TELSTRA":
                            pcount=pcount+1

                if pcount > 0 :
                    listoffields6.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):
                    a=ws['B'+str(i)].value
                    if a is not None and (a.isnumeric()):
                        if ws['G'+str(i)].value != "INSERVICE" and ws['G'+str(i)].value != "CONSTRUCTED":
                            pcount=pcount+1

                if pcount > 0 :
                    listoffields6.append("Life Cycle Status:"+str(pcount))

                
                
            elif ws.title=="Civil - Telstra Ducts" :
                num_rows=ws.max_row
                pcount=0
                for i in range(2,num_rows+1):
                    a=ws['B'+str(i)].value
                    if a is not None and (a.isnumeric()):
                        if (ws['F'+str(i)].value) != "TELSTRA" :
                            pcount=pcount+1

                if pcount > 0 :
                    listoffields7.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):
                    a=ws['B'+str(i)].value
                    if a is not None and (a.isnumeric()):
                        if ws['G'+str(i)].value != "INSERVICE" and ws['G'+str(i)].value != "CONSTRUCTED":
                            pcount=pcount+1

                if pcount > 0 :
                    listoffields7.append("Life Cycle Status:"+str(pcount))

                
                
            elif ws.title=="Civil - Telstra Pits":
                num_rows=ws.max_row
                pcount=0
                for i in range(2,num_rows+1):
                    a=ws['B'+str(i)].value
                    if a is not None and (a.isnumeric()):
                        if (ws['F'+str(i)].value) != "TELSTRA":
                            pcount=pcount+1

                if pcount > 0 :
                    listoffields8.append("Owner Name:"+str(pcount))

                pcount=0
                for i in range(2,num_rows+1):
                    a=ws['B'+str(i)].value
                    if a is not None and (a.isnumeric()):
                        if ws['H'+str(i)].value != "INSERVICE" and ws['H'+str(i)].value != "CONSTRUCTED" :
                            pcount=pcount+1

                if pcount > 0 :
                    listoffields8.append("Life Cycle Status:"+str(pcount))

        textfile = open('activity8.txt', 'w')
        if len(listoffields5) > 0 or len(listoffields4) > 0 or len(listoffields3) > 0 or len(listoffields2) > 0 or len(listoffields1) > 0 :
            d1.update({'Civil - NBN Pits':listoffields1,'Civil - NBN Trenches':listoffields2,'Civil - NBN Ducts':listoffields3,'Plant - Equipment':listoffields4,'Plant - Cables':listoffields5})
            print("Failed")
            status = "Failed\n"
            desc = str(d1)+'\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        else:
            print("Pass")
            status="Pass\n"
            desc="NULL\n"
            textfile.writelines(status)
            textfile.writelines(desc)

        if len(listoffields6) > 0 or len(listoffields7) > 0 or len(listoffields8) > 0:
            d2.update({'Civil - Telstra Trenches':listoffields6,'Civil - Telstra Ducts':listoffields7,'Civil - Telstra Pits':listoffields8})
            print("Failed")
            status = "Failed\n"
            desc = str(d2)+'\n'
            textfile.writelines(status)
            textfile.writelines(desc)
        else:
            print("Pass")
            status = "Pass\n"
            desc = "NULL\n"
            textfile.writelines(status)
            textfile.writelines(desc)
        if len(d1) > 0 or len(d2) > 0 :
            print("Failed")
            status = "Failed\n"
            textfile.writelines(status)

        else:
            print("Success")
            status = "Success\n"
            textfile.writelines(status)

        

if __name__ == "__main__":
    ospob=OSP()
    if os.path.exists('activity8.txt'):
        os.remove('activity8.txt')
    sam=sys.argv[1]
    destdir ='C:\\QI_DATA\\'+ sam
    try:
        files = [f for f in os.listdir(destdir) if os.path.isfile(os.path.join(destdir, f))]
        for f in files:
            if re.search('OSP_Editor_Export_NBN', f):
                osp = 'C:\\QI_DATA\\' + sam + '\\' + f
                break
        ospob.OSPValidation(osp)
    except:
        print("File Improper")
    
